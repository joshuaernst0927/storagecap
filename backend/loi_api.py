from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import Response
from pydantic import BaseModel
from typing import List, Optional
import sys, base64, json, os, shutil
sys.path.insert(0, '/root')
from loi_generator import generate_loi

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# ── LOI ───────────────────────────────────────────────────────────────────────

class LOIData(BaseModel):
    date: str = ""
    property_name: str = ""
    property_address: str = ""
    property_description: str = ""
    asset_type: str = "Self-Storage"
    units: str = ""
    sf: str = ""
    year_built: str = ""
    occupancy: str = ""
    broker1_name: str = ""
    broker2_name: str = ""
    brokerage: str = ""
    broker1_phone: str = ""
    broker2_phone: str = ""
    buyer_broker: str = ""
    salutation: str = ""
    offer_price: str = ""
    all_in_cost: str = ""
    bridge_loan: str = ""
    bridge_rate: str = ""
    sofr: str = ""
    annual_ds: str = ""
    interest_reserve: str = ""
    capex_reserve: str = ""
    gp_fee_total: str = ""
    gp_fee_income: str = ""
    gp_coinvest: str = ""
    lp_equity: str = ""
    going_in_cap: str = ""
    yr3_cap: str = ""
    pf_cap: str = ""
    lp_moic: str = ""
    lp_irr: str = ""
    gp_moic: str = ""
    gp_irr: str = ""
    waterfall: str = ""
    emd: str = ""
    dd_days: str = "45"
    closing_days: str = "45"
    offer_expiry: str = ""
    underwriting_narrative: str = ""
    rent_strategy: str = "25% off first 3 months for new tenants, then full list rate, then 3% every 6 months"
    breakeven_occ: str = "85%"
    exit_cap: str = "7.0%"

@app.post("/generate-loi")
async def generate(data: LOIData):
    pdf_bytes = generate_loi(data.dict())
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=YEM_LOI.pdf"}
    )

@app.get("/health")
async def health():
    return {"status": "ok"}

# ── Run Model (Excel engine) ──────────────────────────────────────────────────

class UnderwriteInputs(BaseModel):
    # Core acquisition
    purchasePrice: float = None
    purchase_price: float = None
    closingCostsPct: float = None
    closing_costs_pct: float = None
    initialRepairs: float = None
    initial_repairs: float = None
    acquisitionFeePct: float = None
    acquisition_fee_pct: float = None
    assetMgmtFeePct: float = None
    dispositionFeePct: float = None
    sponsorCoInvestPct: float = None
    # Operations
    startOccupancy: float = None
    start_occupancy: float = None
    stabilizedOccupancy: float = None
    stabilized_occupancy: float = None
    monthsToStabilization: float = None
    months_to_stabilize: float = None
    annualRentGrowth: float = None
    rent_growth: float = None
    opexGrowth: float = None
    # Debt
    initialLTV: float = None
    initial_ltv: float = None
    initialRate: float = None
    initial_rate: float = None
    initialAmortYears: int = None
    ioPeriodMonths: int = None
    minDSCR: float = None
    # Refi
    refiMonth: int = None
    refiLTV: float = None
    refi_ltv: float = None
    refiRate: float = None
    refi_rate: float = None
    refiAmortYears: int = None
    # Exit
    exitCapRate: float = None
    exit_cap_rate: float = None
    exitMonth: int = None
    exit_month: float = None
    sellingCostsPct: float = None
    # Waterfall
    lpReturnOfCapital: float = None
    lpCatchUp: float = None
    gpCatchUp: float = None
    preferredReturn: float = None
    lpResidual: float = None
    gpResidual: float = None
    unlevered: str = None
    # T12 expenses
    t12Tax: float = None
    t12Insurance: float = None
    t12Utilities: float = None
    t12RepairsMaintenance: float = None
    t12Payroll: float = None
    t12OfficeEmployee: float = None
    t12Marketing: float = None
    t12Administrative: float = None
    # Unit mix
    unitMix: list = None

@app.post("/run-model")
async def run_underwrite(data: UnderwriteInputs):
    inputs = {k: v for k, v in data.dict().items() if v is not None}
    # Normalize snake_case aliases to camelCase for underwrite.py
    aliases = {
        'purchase_price':       'purchasePrice',
        'closing_costs_pct':    'closingCostsPct',
        'initial_repairs':      'initialRepairs',
        'acquisition_fee_pct':  'acquisitionFeePct',
        'start_occupancy':      'startOccupancy',
        'stabilized_occupancy': 'stabilizedOccupancy',
        'months_to_stabilize':  'monthsToStabilization',
        'rent_growth':          'annualRentGrowth',
        'initial_ltv':          'initialLTV',
        'initial_rate':         'initialRate',
        'refi_ltv':             'refiLTV',
        'refi_rate':            'refiRate',
        'exit_cap_rate':        'exitCapRate',
        'exit_month':           'exitMonth',
    }
    for old_key, new_key in aliases.items():
        if old_key in inputs and new_key not in inputs:
            inputs[new_key] = inputs.pop(old_key)

    from underwrite import run_model
    results = run_model(inputs)
    return results

from fastapi.responses import FileResponse
import openpyxl
from datetime import datetime

class DownloadInputs(BaseModel):
    purchasePrice: float = 4250000
    closingCostsPct: float = 0.03
    initialRepairs: float = 150000
    acquisitionFeePct: float = 0.02
    startOccupancy: float = 0.707
    stabilizedOccupancy: float = 0.90
    monthsToStabilization: float = 18
    annualRentGrowth: float = 0.05
    exitCapRate: float = 0.0725
    exitMonth: float = 60
    initialLTV: float = 0.45
    initialRate: float = 0.085
    refiLTV: float = 0.65
    refiRate: float = 0.0525
    unlevered: str = "No"

@app.post("/download-model")
async def download_model(data: DownloadInputs):
    from underwrite import populate
    from pathlib import Path
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    os.makedirs("/root/model_runs", exist_ok=True)
    work_file = f"/root/model_runs/model_{ts}.xlsx"
    populate(data.dict(), Path(work_file))
    return FileResponse(
        work_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="YEM_Acquisition_Model.xlsx",
        background=None
    )

# ── Max Offer ─────────────────────────────────────────────────────────────────

from underwrite import find_max_offer, find_max_offer_by_type

class MaxOfferInputs(BaseModel):
    target_irr: float = 0.15
    deal_type: str = "value-add"
    in_place_noi: float = None
    stabilized_noi: float = None
    purchase_price: float = None
    closing_costs_pct: float = 0.03
    initial_repairs: float = 0
    acquisition_fee_pct: float = 0.02
    start_occupancy: float = 0.75
    stabilized_occupancy: float = 0.90
    months_to_stabilization: float = 18
    rent_growth: float = 0.05
    opex_growth: float = 0.025
    exit_cap_rate: float = 0.0725
    exit_month: float = 60
    selling_costs_pct: float = 0.02
    other_income_month: float = 0

@app.post("/max-offer")
async def max_offer(data: MaxOfferInputs):
    inputs = data.dict()
    target = inputs.pop("target_irr")
    deal_type = inputs.pop("deal_type")
    result = find_max_offer_by_type(inputs, target, deal_type)
    return result

# ── Proforma Builder ──────────────────────────────────────────────────────────

from proforma_engine import build_proforma_from_t12

class ProformaRequest(BaseModel):
    t12_data: dict
    assumptions: dict

@app.post("/build-proforma")
async def build_proforma(req: ProformaRequest):
    result = build_proforma_from_t12(req.t12_data, req.assumptions)
    return result

# ── IRR v2 ────────────────────────────────────────────────────────────────────

def interp_noi_at_month(noi_years, month):
    if not noi_years:
        return 0
    year_exact = month / 12.0
    lower_idx = max(0, int(year_exact) - 1)
    upper_idx = min(len(noi_years) - 1, lower_idx + 1)
    frac = year_exact - int(year_exact)
    lower_noi = noi_years[lower_idx] if lower_idx < len(noi_years) else 0
    upper_noi = noi_years[upper_idx] if upper_idx < len(noi_years) else lower_noi
    if lower_idx == upper_idx or frac == 0:
        return lower_noi
    return lower_noi + frac * (upper_noi - lower_noi)

def calc_loan_balance(original_loan, interest_rate, amort_years, io_months, months_elapsed):
    if months_elapsed <= io_months:
        return original_loan
    amort_months_elapsed = int(months_elapsed - io_months)
    if amort_years <= 0 or interest_rate <= 0:
        return original_loan
    monthly_rate = interest_rate / 12
    n_payments = amort_years * 12
    monthly_payment = original_loan * (monthly_rate * (1 + monthly_rate) ** n_payments) / \
                      ((1 + monthly_rate) ** n_payments - 1)
    balance = original_loan
    for _ in range(amort_months_elapsed):
        interest = balance * monthly_rate
        principal = monthly_payment - interest
        balance -= principal
    return max(0, balance)

def calc_irr_with_debt(purchase_price, noi_years, exit_cap, selling_costs,
                        closing_costs, acq_fee, initial_repairs,
                        ltv, interest_rate, amort_years, io_months,
                        exit_month=None, refi_month=None, refi_ltv=None,
                        refi_rate=None, refi_amort_years=None, refi_dscr=None,
                        refi_fee_pct=None, exit_value_override=None):
    hold_years = len(noi_years)
    hold_months = int(exit_month) if exit_month else hold_years * 12
    exit_year = max(1, min(hold_years, round(hold_months / 12)))
    total_cost = purchase_price * (1 + closing_costs + acq_fee) + initial_repairs
    exit_noi = interp_noi_at_month(noi_years, hold_months)
    if exit_value_override and exit_value_override > 0:
        exit_value = exit_value_override
    else:
        exit_value = exit_noi / exit_cap if exit_cap > 0 else 0
    net_sale = exit_value * (1 - selling_costs)
    ucf = [-total_cost]
    for yr in range(1, exit_year + 1):
        noi = noi_years[yr - 1] if yr <= len(noi_years) else noi_years[-1]
        ucf.append(noi + net_sale if yr == exit_year else noi)
    loan = purchase_price * ltv
    equity = total_cost - loan
    io_years = io_months // 12
    annual_ds_io = loan * interest_rate
    if amort_years > 0 and interest_rate > 0:
        mr = interest_rate / 12
        np_ = amort_years * 12
        mp = loan * (mr * (1 + mr) ** np_) / ((1 + mr) ** np_ - 1)
        annual_ds_bridge = mp * 12
    else:
        annual_ds_bridge = loan * interest_rate
    refi_cash_out = 0
    refi_fee_paid = 0
    new_loan = loan
    new_loan_ds = annual_ds_bridge
    refi_year = None
    rr = interest_rate
    ra = amort_years
    refi_occurs = bool(refi_month and refi_month > 0 and ltv > 0)
    if refi_occurs:
        refi_year = max(1, min(exit_year, round(int(refi_month) / 12)))
        refi_noi = interp_noi_at_month(noi_years, int(refi_month))
        rr = refi_rate if refi_rate else interest_rate
        ra = int(refi_amort_years) if refi_amort_years else 30
        rc = refi_ltv if refi_ltv else 0.70
        rd = refi_dscr if refi_dscr else 1.30
        rf = refi_fee_pct if refi_fee_pct else 0.01
        going_in_cap = noi_years[0] / purchase_price if purchase_price > 0 else 0.07
        refi_stab_value = refi_noi / going_in_cap if going_in_cap > 0 else 0
        ltv_max = refi_stab_value * rc
        dscr_max = (refi_noi / rd / rr) if rr > 0 else 0
        new_loan = min(ltv_max, dscr_max)
        bridge_balance = calc_loan_balance(loan, interest_rate, amort_years, io_months, int(refi_month))
        refi_cash_out = max(0, new_loan - bridge_balance)
        refi_fee_paid = new_loan * rf
        if ra > 0 and rr > 0:
            mr2 = rr / 12
            np2 = ra * 12
            mp2 = new_loan * (mr2 * (1 + mr2) ** np2) / ((1 + mr2) ** np2 - 1)
            new_loan_ds = mp2 * 12
        else:
            new_loan_ds = new_loan * rr
    lcf = [-equity]
    for yr in range(1, exit_year + 1):
        noi = noi_years[yr - 1] if yr <= len(noi_years) else noi_years[-1]
        if refi_occurs and refi_year and yr > refi_year:
            ds = new_loan_ds
        elif yr <= io_years:
            ds = annual_ds_io
        else:
            ds = annual_ds_bridge
        if yr < exit_year:
            cf = noi - ds
            if refi_occurs and refi_year and yr == refi_year:
                cf += refi_cash_out - refi_fee_paid
            lcf.append(cf)
        else:
            if refi_occurs and refi_year:
                months_since_refi = (exit_year - refi_year) * 12
                remaining_balance = calc_loan_balance(new_loan, rr, ra, 0, months_since_refi)
            else:
                remaining_balance = calc_loan_balance(loan, interest_rate, amort_years, io_months, hold_months)
            cf = noi - ds + net_sale - remaining_balance
            if refi_occurs and refi_year and yr == refi_year:
                cf += refi_cash_out - refi_fee_paid
            lcf.append(cf)

    def irr_calc(cfs):
        def npv(r):
            return sum(cf / (1 + r) ** i for i, cf in enumerate(cfs))
        lo, hi = -0.9, 10.0
        for _ in range(200):
            mid = (lo + hi) / 2
            if npv(mid) > 0:
                lo = mid
            else:
                hi = mid
        return mid

    u_irr = irr_calc(ucf)
    l_irr = irr_calc(lcf)
    equity_multiple = sum(cf for cf in lcf if cf > 0) / equity if equity > 0 else 0
    annual_ds_display = annual_ds_bridge if not refi_occurs else new_loan_ds
    return {
        "unlevered_irr":       round(u_irr, 4),
        "levered_irr":         round(l_irr, 4),
        "equity_multiple":     round(equity_multiple, 2),
        "equity_required":     round(equity),
        "loan_amount":         round(loan),
        "annual_debt_service": round(annual_ds_display),
        "going_in_cap":        round(noi_years[0] / purchase_price, 4) if purchase_price > 0 else 0,
        "stabilized_cap":      round(exit_noi / purchase_price, 4) if purchase_price > 0 else 0,
        "exit_value":          round(exit_value),
        "refi_cash_out":       round(refi_cash_out),
        "refi_fee_paid":       round(refi_fee_paid),
        "new_loan":            round(new_loan),
    }

class CalcIRRV2Request(BaseModel):
    purchase_price: float
    noi_years: list
    exit_cap_rate: float = 0.075
    exit_month: int = 60
    exit_value_override: float = None
    selling_costs_pct: float = 0.02
    closing_costs_pct: float = 0.03
    acquisition_fee_pct: float = 0.02
    initial_repairs: float = 0
    ltv: float = 0.65
    interest_rate: float = 0.07
    amort_years: int = 30
    io_months: int = 24
    refi_month: int = None
    refi_ltv: float = None
    refi_rate: float = None
    refi_amort_years: int = None
    refi_dscr: float = None
    refi_fee_pct: float = None

@app.post("/calc-irr-v2")
async def calc_irr_v2(req: CalcIRRV2Request):
    result = calc_irr_with_debt(
        req.purchase_price, req.noi_years, req.exit_cap_rate,
        req.selling_costs_pct, req.closing_costs_pct, req.acquisition_fee_pct,
        req.initial_repairs, req.ltv, req.interest_rate, req.amort_years, req.io_months,
        exit_month=req.exit_month, refi_month=req.refi_month, refi_ltv=req.refi_ltv,
        refi_rate=req.refi_rate, refi_amort_years=req.refi_amort_years,
        refi_dscr=req.refi_dscr, refi_fee_pct=req.refi_fee_pct,
        exit_value_override=req.exit_value_override,
    )
    return result

class CalcIRRRequest(BaseModel):
    purchase_price: float
    in_place_noi: float
    stabilized_noi: float
    start_occupancy: float = 0.75
    stabilized_occupancy: float = 0.92
    exit_cap_rate: float = 0.075
    exit_month: float = 60
    months_to_stabilization: int = 18
    rent_growth: float = 0.03
    opex_growth: float = 0.03
    closing_costs_pct: float = 0.03
    acquisition_fee_pct: float = 0.02
    initial_repairs: float = 0
    selling_costs_pct: float = 0.02

@app.post("/calc-irr")
async def calc_irr_endpoint(req: CalcIRRRequest):
    from underwrite import run_model
    inputs = req.dict()
    result = run_model(inputs)
    return {
        "irr_at_price": result.get("levered_irr", 0),
        "going_in_cap": round(req.in_place_noi / req.purchase_price, 4),
        "stabilized_cap": round(req.stabilized_noi / req.purchase_price, 4),
        "in_place_noi": req.in_place_noi,
        "stabilized_noi": req.stabilized_noi,
        "purchase_price": req.purchase_price,
    }

# ── Document Extraction ───────────────────────────────────────────────────────

EXTRACTION_PROMPT = """You are analyzing self-storage acquisition documents (rent roll, T12 P&L, offering memorandum, proforma, or deal memo).

Extract ALL available inputs for financial underwriting. Return ONLY a valid JSON object - no markdown fences, no commentary, no extra text. Use null for any field you cannot find.

CRITICAL - For Excel/spreadsheet files, look carefully for:
- Unit mix tables (rows with unit sizes like 5x5, 5x10, 10x10, 10x15, 10x20 and their counts, rents)
- Rent roll summaries showing total units, occupied units, current rents
- T12 or trailing 12-month income statements with ALL expense line items
- Proforma projections labeled Year 1, Year 2, Year 3, Y1, Y2, Y3

IMPORTANT: Extract the SELLER'S projected numbers exactly as presented. Do not adjust or haircut them.

Extract these fields:
- propertyName, address, city, state, dealType (value-add/stabilized/distressed)
- totalUnits, currentOccupancy (percent e.g. 85), currentAvgRentPerUnit (CURRENT rent tenants actually pay today), marketAvgRentPerUnit (MARKET rate from comparables - must be DIFFERENT from current rent for value-add deals)
- t12Revenue, t12Expenses, t12NOI, t3NOI
- t12Payroll, t12ManagementFees, t12Marketing, t12Utilities
- t12OfficeEmployee, t12Administrative, t12RepairsMaintenance
- t12Tax, t12Insurance, t12OtherExpenses
- sellerY1Revenue, sellerY1Expenses, sellerY1NOI
- sellerY2Revenue, sellerY2Expenses, sellerY2NOI
- sellerY3Revenue, sellerY3Expenses, sellerY3NOI
- sellerY4Revenue, sellerY4NOI, sellerY5Revenue, sellerY5NOI
- monthsToStabilization, projectedStabilizedOccupancy, projectedStabilizedNOI
- purchasePrice, exitCapRate, exitMonth
- yearBuilt, sqft (total net rentable SF)
- broker1Name, broker2Name, brokerPhone1, brokerPhone2, brokerEmail1, brokerEmail2, brokerageName
- unitMix: array of {type, units, sqft, currentRent, marketRent}

Return JSON only."""

from fastapi import UploadFile
import mimetypes, io

@app.post("/extract")
async def extract_documents(files: list[UploadFile]):
    import anthropic
    client = anthropic.Anthropic()
    content_blocks = []
    for upload in files:
        raw = await upload.read()
        label = upload.filename or "document"
        mime = upload.content_type or mimetypes.guess_type(label)[0] or "application/octet-stream"
        b64 = base64.b64encode(raw).decode()
        if mime == "application/pdf":
            content_blocks.append({"type": "text", "text": f"--- File: {label} ---"})
            content_blocks.append({"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": b64}})
        elif mime.startswith("image/"):
            content_blocks.append({"type": "text", "text": f"--- File: {label} ---"})
            content_blocks.append({"type": "image", "source": {"type": "base64", "media_type": mime, "data": b64}})
        elif "spreadsheetml" in mime or label.lower().endswith(".xlsx"):
            wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True)
            text_parts = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                for row in ws.iter_rows(values_only=True):
                    row_vals = [str(c) if c is not None else "" for c in row]
                    if any(v.strip() for v in row_vals):
                        rows.append(",".join(row_vals))
                if rows:
                    text_parts.append(f"Sheet: {sheet_name}\n" + "\n".join(rows))
            text = "\n\n".join(text_parts)[:30000]
            content_blocks.append({"type": "text", "text": f"--- File: {label} ---\n\n{text}"})
        else:
            try:
                text = raw.decode("utf-8", errors="replace")[:25000]
                content_blocks.append({"type": "text", "text": f"--- File: {label} ---\n\n{text}"})
            except Exception:
                pass
    content_blocks.append({"type": "text", "text": EXTRACTION_PROMPT})
    msg = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=3000,
        messages=[{"role": "user", "content": content_blocks}]
    )
    raw_text = msg.content[0].text.strip()
    if "```json" in raw_text:
        raw_text = raw_text.split("```json")[1].split("```")[0].strip()
    elif raw_text.startswith("```"):
        raw_text = raw_text.split("\n", 1)[1] if "\n" in raw_text else raw_text
        raw_text = raw_text.rsplit("```", 1)[0].strip()
    if not raw_text:
        return {"error": "empty response from Claude", "raw": str(msg.content)}
    try:
        return json.loads(raw_text)
    except Exception as e:
        return {"error": str(e), "raw": raw_text[:500]}
